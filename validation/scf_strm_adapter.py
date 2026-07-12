"""
scf_strm_adapter.py - Parse SCF Set Theory Relationship Mapping (STRM) workbooks.

STRM (NIST IR 8477) records, per (focal-document element, SCF control), a typed
set relationship and a strength rating - far richer than a flat co-mapping
sheet. SCF ships one STRM workbook per framework. Header is on row 5; each
data row:
  FDE # | FDE Name | FDE Description | STRM Rationale | STRM Relationship |
  SCF Control | SCF # | SCF Description | Strength of Relationship | Notes

Relationship types: Equal, Subset Of, Superset Of, Intersects With, No Relationship.
Strength of Relationship (SoR): ~3-10, higher = stronger.

Output: out/scf_strm.json = {framework: {scf_id: [{ref, name, rel, sor}]}}.
SCF is CC BY-ND - output kept local (gitignored). See validation/README.md.

Frameworks and their STRM workbook filenames are supplied on the command line
via repeatable --map FRAMEWORK=filename.xlsx pairs, so this script carries no
knowledge of any particular deployment's framework set.

Refs are normalised per framework using one of four built-in styles, chosen
with --ref-style FRAMEWORK=STYLE (repeatable; defaults to "iso" for any
framework not given a style):
  iso     - dotted numeric refs, e.g. "5.17" or "A.5.17"
  caf     - lettered-outcome refs, e.g. "A1.a"
  article - "Article N(...)" refs (e.g. GDPR), sub-clauses stripped
  section - "Section N.N" refs

Run:
    python validation/scf_strm_adapter.py \
        --dir path/to/strm/workbooks \
        --map ISO27001=scf-strm-general-iso-27001-2022.xlsx \
        --map ISO27002=scf-strm-general-iso-27002-2022.xlsx \
        --map CAF=scf-strm-emea-gbr-caf-4-0.xlsx \
        --ref-style CAF=caf \
        --out validation/out/scf_strm.json
"""

from __future__ import annotations

import argparse
import json
import re
from pathlib import Path

import openpyxl

HERE = Path(__file__).parent


def norm(s) -> str:
    return re.sub(r"\s+", " ", str(s or "")).strip()


def strip_parens(s: str) -> str:
    """Drop parenthetical annotations, e.g. 'Article 5(1)(a)' -> 'Article 5'."""
    return re.sub(r"\([^)]*\)", "", norm(s)).strip()


# --- ref-normalisation styles -----------------------------------------------

def ref_norm_iso(s: str) -> str:
    """ISO-style dotted numeric refs: 'A.5.17 Title text' -> 'A.5.17'."""
    s = strip_parens(s)
    m = re.match(r"^[A-Za-z]?\.?\d+(?:\.\d+)*", s)
    return m.group(0) if m else s


def ref_norm_caf(s: str) -> str:
    """CAF-style lettered outcomes: 'A1.a Some title' -> 'A1.a'."""
    s = strip_parens(s)
    m = re.match(r"^[A-Za-z]+\d+(?:\.[A-Za-z0-9]+)*", s)
    return m.group(0) if m else s


def ref_norm_article(s: str) -> str:
    """Article-style refs (e.g. GDPR): 'Article 5(1)(a)' -> 'Article 5'."""
    s = strip_parens(s)
    m = re.match(r"^(Article\s+\d+)", s, re.IGNORECASE)
    return m.group(1) if m else s


def ref_norm_section(s: str) -> str:
    """Section-style refs: 'Section 2.3 - Access control' -> 'Section 2.3'."""
    s = strip_parens(s)
    m = re.match(r"^(Section\s+\d+(?:\.\d+)*)", s, re.IGNORECASE)
    return m.group(1) if m else s


REF_STYLES = {
    "iso": ref_norm_iso,
    "caf": ref_norm_caf,
    "article": ref_norm_article,
    "section": ref_norm_section,
}
DEFAULT_REF_STYLE = "iso"


def col_index(header: list[str]) -> dict[str, int]:
    idx = {}
    for i, h in enumerate(header):
        h = norm(h)
        if h == "FDE #":
            idx["ref"] = i
        elif h == "FDE Name":
            idx["name"] = i
        elif h == "STRM Relationship":
            idx["rel"] = i
        elif h == "Strength of Relationship":
            idx["sor"] = i
        elif h == "SCF #":
            idx["scf"] = i
    return idx


def parse_kv_args(pairs: list[str], flag: str) -> dict[str, str]:
    out = {}
    for p in pairs:
        if "=" not in p:
            raise SystemExit(f"{flag} expects FRAMEWORK=VALUE, got: {p!r}")
        k, v = p.split("=", 1)
        out[k.strip()] = v.strip()
    return out


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--dir", required=True, help="directory containing STRM workbooks")
    ap.add_argument("--map", action="append", default=[], required=True,
                     metavar="FRAMEWORK=filename.xlsx",
                     help="repeatable: framework -> STRM workbook filename")
    ap.add_argument("--ref-style", action="append", default=[],
                     metavar="FRAMEWORK=iso|caf|article|section",
                     help="repeatable: ref-normalisation style per framework "
                          f"(default: {DEFAULT_REF_STYLE})")
    ap.add_argument("--out", default=str(HERE / "out" / "scf_strm.json"))
    args = ap.parse_args()

    files = parse_kv_args(args.map, "--map")
    styles_raw = parse_kv_args(args.ref_style, "--ref-style")
    for fw, style in styles_raw.items():
        if style not in REF_STYLES:
            raise SystemExit(f"--ref-style {fw}={style}: unknown style, "
                              f"choose one of {sorted(REF_STYLES)}")

    d = Path(args.dir)
    out: dict[str, dict[str, list]] = {}
    totals = {}
    for fw, fname in files.items():
        ref_norm = REF_STYLES[styles_raw.get(fw, DEFAULT_REF_STYLE)]
        wb = openpyxl.load_workbook(d / fname, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(min_row=5, values_only=True))
        idx = col_index([norm(c) for c in rows[0]])
        by_scf: dict[str, list] = {}
        n = 0
        for r in rows[1:]:
            rel = norm(r[idx["rel"]])
            scf = norm(r[idx["scf"]])
            if rel in ("", "No Relationship") or scf in ("", "N/A"):
                continue
            ref = ref_norm(r[idx["ref"]])
            sor_raw = norm(r[idx["sor"]])
            sor = int(sor_raw) if sor_raw.isdigit() else None
            by_scf.setdefault(scf, []).append(
                {"ref": ref, "name": norm(r[idx["name"]]), "rel": rel, "sor": sor})
            n += 1
        out[fw] = by_scf
        totals[fw] = n

    outp = Path(args.out)
    outp.parent.mkdir(parents=True, exist_ok=True)
    outp.write_text(json.dumps(out, ensure_ascii=False), encoding="utf-8")
    print("STRM legs parsed (rel != No Relationship):", totals)
    print(f"Wrote {outp}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
